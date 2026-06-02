"""
build_model.py — Extract a friendly input/output *schema* from the CA-GREET4
Tier 1 Dairy/Swine Manure biomethane calculator and serialize it to model.pkl.

The schema is what the Streamlit app reads to render its UI. We extract:
  * setup fields   -> dropdowns and single-value scalar inputs
  * monthly tables -> the wide metered-data grids (auto-detected editable cols)
  * outputs        -> the CNG / LNG / L-CNG carbon-intensity results
  * breakdown      -> the CI contribution rows for the stacked chart

Run once whenever the source workbook changes:
    python app/build_model.py
"""
from __future__ import annotations

import pickle
import warnings
from dataclasses import dataclass, field, asdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = "t1_biomethane_ad_dairy_swine_manure_simplified_calculator_07012025_2.xlsm"
INPUT_FILL = "FFFFFFCC"  # light-yellow "enter input here" highlight used in the sheet
OUT_PATH = Path(__file__).resolve().parent / "model.pkl"


# --------------------------------------------------------------------------- #
# Schema dataclasses
# --------------------------------------------------------------------------- #
@dataclass
class Field:
    key: tuple            # (SHEET_UPPER, "A1")
    label: str
    group: str
    type: str = "number"  # number | text | select
    options: list = field(default_factory=list)
    default: object = None
    unit: str = ""
    help: str = ""


@dataclass
class Column:
    col: str              # column letter
    label: str
    type: str = "number"
    options: list = field(default_factory=list)
    unit: str = ""


@dataclass
class Table:
    id: str
    title: str
    sheet: str            # UPPER-cased sheet name (engine key form)
    row_start: int
    row_end: int
    index_col: str
    index_label: str
    columns: list = field(default_factory=list)
    scalars: list = field(default_factory=list)  # per-table single-cell inputs
    note: str = ""


@dataclass
class Output:
    key: tuple
    label: str
    unit: str = ""


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def is_input(cell) -> bool:
    return bool(cell.fill and cell.fill.patternType and cell.fill.fgColor.rgb == INPUT_FILL)


def clean(v) -> str:
    return "" if v is None else str(v).replace("\n", " ").strip()


def dv_options(ws, coord):
    """Return inline dropdown options for a cell, if any."""
    if not ws.data_validations:
        return None
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and dv.formula1 and coord in dv.sqref:
            f1 = dv.formula1.strip().strip('"')
            return [o.strip() for o in f1.split(",")]
    return None


def header_label(ws, col_idx, header_rows, unit_row=None):
    """Build a column label from header text row(s); optional unit row."""
    label = ""
    for r in header_rows:
        t = clean(ws.cell(r, col_idx).value)
        if t and not t.startswith("="):
            label = t
            break
    unit = ""
    if unit_row:
        unit = clean(ws.cell(unit_row, col_idx).value)
    return label, unit


def left_label(ws, cell):
    for col in range(cell.column - 1, 0, -1):
        v = ws.cell(cell.row, col).value
        if isinstance(v, str) and v.strip() and not v.startswith("="):
            return v.replace("\n", " ").strip()
    return ""


# --------------------------------------------------------------------------- #
# Extraction
# --------------------------------------------------------------------------- #
def extract_production_table(ws) -> Table:
    """Section 2 monthly metered data on Biogas-to-RNG (rows 28-51)."""
    r0, r1 = 28, 51
    cols = []
    for ci in range(column_index_from_string("C"), column_index_from_string("AH") + 1):
        letter = get_column_letter(ci)
        if not is_input(ws.cell(r0, ci)):
            continue
        label, _ = header_label(ws, ci, [26])
        unit = clean(ws.cell(27, ci).value)
        opts = dv_options(ws, f"{letter}{r0}")
        cols.append(Column(col=letter, label=label or letter,
                           type="select" if opts else "number",
                           options=opts or [], unit=unit))
    return Table(
        id="production",
        title="Section 2 — Monthly Biomethane Production Data",
        sheet="BIOGAS-TO-RNG",
        row_start=r0, row_end=r1,
        index_col="B", index_label="Reporting Month (MM/YYYY)",
        columns=cols,
        note="Metered monthly data for raw biogas, upgrading, energy use, "
             "biomethane injected, and transport. Enter one row per reporting month.",
    )


def extract_manure_tables(ws, livestock_opts) -> list:
    """Section L1 baseline tables — up to 6 livestock-category blocks (14 cols each)."""
    tables = []
    r0, r1 = 9, 32
    for i in range(6):
        base = 2 + 14 * i           # B, P, AD, AR, BF, BT ...
        base_letter = get_column_letter(base)
        index_col_idx = base + 1    # month column (C, Q, ...)
        cat_name = clean(ws.cell(6, base).value) or f"Livestock Category {i + 1}"
        # per-block scalar inputs: livestock category (row 7) + baseline reporting period (row 9)
        scalars = [
            {"cell": f"{base_letter}7", "label": "Livestock Category",
             "type": "select", "options": livestock_opts, "default": livestock_opts[0]},
            {"cell": f"{base_letter}9", "label": "Baseline Reporting Period (months)",
             "type": "select", "options": ["12", "24", "-"], "default": "12"},
        ]
        cols = []
        for off in range(1, 13):    # within-block columns
            ci = base + off
            letter = get_column_letter(ci)
            # editable if any data row in this column is yellow
            if not any(is_input(ws.cell(r, ci)) for r in range(r0, r1 + 1)):
                continue
            label, _ = header_label(ws, ci, [6])
            unit = clean(ws.cell(7, ci).value)
            opts = dv_options(ws, f"{letter}{r0}")
            cols.append(Column(col=letter, label=label or letter,
                               type="select" if opts else "number",
                               options=opts or [], unit=unit))
        if not cols:
            continue
        tables.append(Table(
            id=f"manure_cat{i + 1}",
            title=f"Baseline Methane — {cat_name}",
            sheet="MANURE-TO-BIOGAS (LOP INPUTS)",
            row_start=r0, row_end=r1,
            index_col=get_column_letter(index_col_idx),
            index_label="Reporting Month (MM/YYYY)",
            columns=cols,
            scalars=scalars,
            note="Per-category monthly herd data feeding the baseline anaerobic-"
                 "storage methane calculation.",
        ))
    return tables


def build():
    path = ROOT / WORKBOOK
    wb = openpyxl.load_workbook(path, keep_vba=True)
    wb_vals = openpyxl.load_workbook(path, data_only=True)  # cached values for defaults

    rng = wb["Biogas-to-RNG"]
    ae = wb["Avoided Emissions"]
    ef = wb["EF Table"]

    # --- electricity-mix options from EF Table (D6:D33) ------------------ #
    elec_opts = ["User Defined Mix"]
    for r in range(6, 34):
        name = clean(ef.cell(r, column_index_from_string("D")).value)
        if name:
            elec_opts.append(name)

    # --- livestock categories from Reference A94:A103 -------------------- #
    refsheet = wb["Reference"]
    livestock_opts = []
    for r in range(94, 104):
        name = clean(refsheet.cell(r, 1).value)
        if name and name != "-":
            livestock_opts.append(name)

    # --- setup fields (curated, high-level toggles) --------------------- #
    def default(sheet_vals, coord):
        return sheet_vals[coord].value

    def num_default(sheet_vals, coord):
        v = sheet_vals[coord].value
        return v if isinstance(v, (int, float)) else None

    setup = [
        Field(key=("BIOGAS-TO-RNG", "C15"), label="Company Name and ID", group="Applicant Info",
              type="text", default=default(wb_vals["Biogas-to-RNG"], "C15"), help="1.1"),
        Field(key=("BIOGAS-TO-RNG", "C16"), label="Facility Name and ID", group="Applicant Info",
              type="text", default=default(wb_vals["Biogas-to-RNG"], "C16"), help="1.2"),
        Field(key=("BIOGAS-TO-RNG", "C18"), label="Digester Location (Street, City)", group="Applicant Info",
              type="text", default=default(wb_vals["Biogas-to-RNG"], "C18"), help="1.4.a"),
        Field(key=("BIOGAS-TO-RNG", "C19"), label="Digester Location (State)", group="Applicant Info",
              type="text", default=default(wb_vals["Biogas-to-RNG"], "C19") or "California", help="1.4.b"),

        Field(key=("AVOIDED EMISSIONS", "C4"), label="Digester Type", group="Project Setup",
              type="select", options=dv_options(ae, "C4") or ["Covered Lagoon", "Enclosed Vessel"],
              default=default(wb_vals["Avoided Emissions"], "C4") or "Covered Lagoon", help="P1.1"),
        Field(key=("BIOGAS-TO-RNG", "D24"), label="Electricity Mix for Biomethane", group="Project Setup",
              type="select", options=elec_opts,
              default=default(wb_vals["Biogas-to-RNG"], "D24") or "CAMX", help="2.1"),
        Field(key=("BIOGAS-TO-RNG", "AH34"), label="Pipeline Distance: Upgrading → CNG Station", group="Project Setup",
              type="number", unit="miles", default=num_default(wb_vals["Biogas-to-RNG"], "AH34"), help="2.35.a"),
        Field(key=("BIOGAS-TO-RNG", "AH40"), label="Pipeline Distance: Upgrading → LNG Plant", group="Project Setup",
              type="number", unit="miles", default=num_default(wb_vals["Biogas-to-RNG"], "AH40"), help="2.35.b"),
        Field(key=("BIOGAS-TO-RNG", "AH54"), label="Fugitive Methane from Upgrading", group="Project Setup",
              type="number", unit="fraction 0–1", default=0.02, help="2.37"),

        Field(key=("BIOGAS-TO-RNG", "AL24"), label="LNG / L-CNG Facility Name and ID", group="LNG / L-CNG",
              type="text", default=default(wb_vals["Biogas-to-RNG"], "AL24"), help="3.1"),
        Field(key=("BIOGAS-TO-RNG", "AL25"), label="Liquefaction → Station Truck Distance", group="LNG / L-CNG",
              type="number", unit="miles", default=num_default(wb_vals["Biogas-to-RNG"], "AL25"), help="3.2"),
        Field(key=("BIOGAS-TO-RNG", "AL26"), label="Liquefaction Emission Factor", group="LNG / L-CNG",
              type="number", unit="gCO2e/gal LNG", default=num_default(wb_vals["Biogas-to-RNG"], "AL26"), help="3.3"),
        Field(key=("BIOGAS-TO-RNG", "AL27"), label="LNG Truck Type", group="LNG / L-CNG",
              type="select", options=dv_options(rng, "AL27") or ["Boil-Off Recovery Equipped", "No Boil-Off Recovery"],
              default=default(wb_vals["Biogas-to-RNG"], "AL27") or "Boil-Off Recovery Equipped", help="3.4"),
    ]

    # --- tables --------------------------------------------------------- #
    tables = [extract_production_table(rng)] + extract_manure_tables(
        wb["Manure-to-Biogas (LOP Inputs)"], livestock_opts)

    # --- outputs & breakdown ------------------------------------------- #
    outputs = [
        Output(key=("PATHWAY SUMMARY", "F59"), label="CNG Carbon Intensity", unit="gCO2e/MJ"),
        Output(key=("PATHWAY SUMMARY", "F67"), label="LNG Carbon Intensity", unit="gCO2e/MJ"),
        Output(key=("PATHWAY SUMMARY", "F75"), label="L-CNG Carbon Intensity", unit="gCO2e/MJ"),
    ]
    breakdown = [
        Output(key=("PATHWAY SUMMARY", "F28"), label="Biogas Production & Upgrading"),
        Output(key=("PATHWAY SUMMARY", "F31"), label="Post-Digester Fugitives"),
        Output(key=("PATHWAY SUMMARY", "F32"), label="Flared Biomethane"),
        Output(key=("PATHWAY SUMMARY", "F35"), label="Biomethane Transmission"),
        Output(key=("PATHWAY SUMMARY", "F48"), label="CNG Production"),
        Output(key=("PATHWAY SUMMARY", "F49"), label="Tailpipe Emissions"),
        Output(key=("PATHWAY SUMMARY", "F21"), label="Avoided / Diverted Credits"),
    ]

    model = {
        "workbook": WORKBOOK,
        "title": "CA-GREET4 Tier 1 — Dairy / Swine Manure Biomethane",
        "setup_fields": [asdict(f) for f in setup],
        "tables": [asdict(t) for t in tables],
        "outputs": [asdict(o) for o in outputs],
        "breakdown": [asdict(o) for o in breakdown],
    }

    with open(OUT_PATH, "wb") as fh:
        pickle.dump(model, fh)

    print(f"Wrote {OUT_PATH}")
    print(f"  setup fields : {len(setup)}")
    print(f"  tables       : {len(tables)}")
    for t in tables:
        print(f"     - {t.id:12s} {len(t.columns):2d} cols  rows {t.row_start}-{t.row_end}")
    print(f"  outputs      : {len(outputs)}  breakdown rows: {len(breakdown)}")


if __name__ == "__main__":
    build()
